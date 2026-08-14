"""
Aegis Node — Rule-Based Content Threat Inspector.
Scans dataset content (CSV, JSON, JSONL, Parquet, XLSX, TXT) for known injection patterns.

Security note:
  - Files are READ ONLY via pandas/json/openpyxl parsers.
  - No user-supplied content is eval()d or exec()d.
  - Only first 10,000 rows are inspected to bound memory usage.
  - All cell values are deobfuscated before rule matching (URL decode, HTML unescape, SQL comment removal).
"""

import html
import io
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import unquote

import pandas as pd

logger = logging.getLogger(__name__)

# ─── Maximum rows inspected per dataset ─────────────────────────────────────
_MAX_ROWS = 10_000
_XLSX_MAX_MB = 50  # Hard cap for Excel files (openpyxl loads full sheet)
_TXT_MAX_LINES = 10_000
_SAMPLE_MAX_LEN = 200  # max chars shown in finding sample

# ─── Threat Detection Rules ──────────────────────────────────────────────────
# Each rule: (rule_id, severity, category, description, compiled_regex)

_EICAR = b"X5O!P%@AP[4\\PZX54(P^)7CC)7}$EICAR-STANDARD-ANTIVIRUS-TEST-FILE!$H+H*"

# ─── Threat Detection Rules ──────────────────────────────────────────────────
# Each rule: (rule_id, severity, category, description, compiled_regex)

_RULES: list[tuple[str, str, str, str, re.Pattern]] = [
    # ── CSV Formula Injection (OWASP) ──────────────────────────────────────────
    (
        "FORM-001", "high", "formula_injection",
        "CSV formula injection — cell starts with Excel formula trigger character",
        re.compile(r'^\s*[=+\-@|]', re.MULTILINE),
    ),
    (
        "FORM-002", "critical", "formula_injection",
        "DDE command execution payload detected in cell value",
        re.compile(r'DDE\s*\(|cmd\s*\||powershell\s*\|', re.IGNORECASE),
    ),
    (
        "FORM-003", "high", "formula_injection",
        "HYPERLINK formula with external URL detected",
        re.compile(r'HYPERLINK\s*\(', re.IGNORECASE),
    ),
    # ── Script Injection ───────────────────────────────────────────────────────
    (
        "SCRP-001", "critical", "script_injection",
        "HTML/JavaScript <script> tag found inside dataset field",
        re.compile(r'<\s*script[\s>]', re.IGNORECASE),
    ),
    (
        "SCRP-002", "high", "script_injection",
        "javascript: protocol handler found in field value",
        re.compile(r'javascript\s*:', re.IGNORECASE),
    ),
    (
        "SCRP-003", "high", "script_injection",
        "eval() or exec() call found in field value",
        re.compile(r'\beval\s*\(|\bexec\s*\(', re.IGNORECASE),
    ),
    # ── SQL Injection ──────────────────────────────────────────────────────────
    (
        "SQLI-001", "high", "sql_injection",
        "Classic SQL injection payload detected",
        re.compile(r"'\s*OR\s*'1'\s*=\s*'1|--\s|;\s*DROP\s+TABLE", re.IGNORECASE),
    ),
    (
        "SQLI-002", "high", "sql_injection",
        "UNION SELECT injection pattern detected",
        re.compile(r"UNION\s+(ALL\s+)?SELECT", re.IGNORECASE),
    ),
    # ── Binary Anomaly ─────────────────────────────────────────────────────────
    (
        "BIN-001", "medium", "binary_anomaly",
        "Null byte (\\x00) found in text field — possible binary data injection",
        re.compile(r'\x00'),
    ),
    # ── EICAR Test String ─────────────────────────────────────────────────────
    (
        "MAL-001", "critical", "malware_signature",
        "EICAR antivirus test string detected — confirms AV detection pipeline is working",
        re.compile(r'X5O!P%@AP\[4\\\\PZX54\(P\^\)7CC\)7\}', re.IGNORECASE),
    ),
    # ── Executable / Binary Headers in Text Fields ────────────────────────────
    (
        "MAL-002", "critical", "malware_signature",
        "Windows PE/MZ executable header detected in field value",
        re.compile(r'\bMZ\x90|\bMZP\x00|\b4d5a90|\b4d5a50', re.IGNORECASE),
    ),
    (
        "MAL-003", "critical", "malware_signature",
        "ELF binary header detected in field value (Linux executable)",
        re.compile(r'\x7fELF|7f454c46', re.IGNORECASE),
    ),
    (
        "MAL-004", "critical", "malware_signature",
        "Base64-encoded Windows PE executable detected (TVoA or TVJQ prefix = MZ header)",
        re.compile(r'\bTVo[Aqw]|\bTVJQ|\bTVpQ', re.IGNORECASE),
    ),
    # ── PowerShell / Shellcode ────────────────────────────────────────────────
    (
        "MAL-005", "critical", "shellcode",
        "PowerShell encoded command (Base64) detected — common malware dropper technique",
        re.compile(r'powershell\s.*-[Ee]nc(?:odedCommand)?\s+[A-Za-z0-9+/=]{20,}', re.IGNORECASE),
    ),
    (
        "MAL-006", "critical", "shellcode",
        "PowerShell download-and-execute cradle detected",
        re.compile(r'IEX\s*\(|Invoke-Expression\s*\(|DownloadString\s*\(|WebClient\(\)', re.IGNORECASE),
    ),
    (
        "MAL-007", "high", "shellcode",
        "Reverse shell payload pattern detected",
        re.compile(r'bash\s+-i\s+>&\s*/dev/tcp|nc\s+-e\s*/bin/|/bin/sh\s+-i', re.IGNORECASE),
    ),
    # ── Suspicious Macro / OLE patterns ──────────────────────────────────────
    (
        "MAL-008", "high", "macro_threat",
        "Auto-execute macro trigger keyword detected (AutoOpen, Document_Open, etc.)",
        re.compile(r'\b(AutoOpen|Document_Open|Workbook_Open|Auto_Open|Shell|CreateObject|WScript\.Shell)\b', re.IGNORECASE),
    ),
    # ── Known Malware Family Names in Filenames / Fields ─────────────────────
    (
        "MAL-009", "critical", "malware_reference",
        "Known malware family name detected in field — dataset contains malware analysis data",
        re.compile(
            r'\b(Trojan|Ransomware|Spyware|Adware|Rootkit|Backdoor|Worm|Virus|Keylogger|'
            r'Botnet|RAT|Dropper|Downloader|Infostealer|Cryptominer|Fileless|'
            r'Mirai|WannaCry|Petya|NotPetya|Emotet|TrickBot|Ryuk|Conti|'
            r'BlackCat|ALPHV|LockBit|REvil|Sodinokibi|DarkSide|Maze|'
            r'Stuxnet|Duqu|Flame|Carbanak|APT|TIBS|ZeuS|Conficker|'
            r'Dridex|Ursnif|Qakbot|IcedID|BazarLoader)\b',
            re.IGNORECASE,
        ),
    ),
    # ── Suspicious URLs / C2 Patterns ────────────────────────────────────────
    (
        "MAL-010", "high", "c2_communication",
        "Suspicious URL with IP address and port (potential C2 server)",
        re.compile(r'https?://\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}:\d{2,5}', re.IGNORECASE),
    ),
    (
        "MAL-011", "high", "c2_communication",
        "Hex-encoded shellcode pattern detected (\\x41\\x42 style sequences)",
        re.compile(r'(\\x[0-9a-fA-F]{2}){6,}', re.IGNORECASE),
    ),
]


# Severity → numeric weight for risk score calculation
_SEVERITY_WEIGHT = {"critical": 3.5, "high": 2.0, "medium": 1.0, "low": 0.3}


@dataclass
class ContentFinding:
    rule_id: str
    severity: str
    category: str
    description: str
    location: str
    sample: str = ""


@dataclass
class ContentCheckResult:
    findings: list[ContentFinding] = field(default_factory=list)
    rows_inspected: int = 0
    error: str | None = None

    @property
    def threat_count(self) -> int:
        return len(self.findings)

    @property
    def risk_score(self) -> float:
        """Returns a capped 0–10 risk score based on finding severities."""
        total = sum(_SEVERITY_WEIGHT.get(f.severity, 0) for f in self.findings)
        return round(min(total, 10.0), 2)


def _deobfuscate(value: str) -> str:
    """
    Multi-stage deobfuscation preprocessor applied before rule matching.
    Normalises common encoding tricks used to evade regex detection.
    Security: only decodes – never executes any content.
    """
    # 1. URL decode (handles %3D, %27, etc.)
    try:
        value = unquote(value)
    except Exception:  # noqa: BLE001
        pass
    # 2. HTML entity unescape (&lt; → <, &#39; → ', etc.)
    try:
        value = html.unescape(value)
    except Exception:  # noqa: BLE001
        pass
    # 3. Strip SQL inline comments /* ... */
    value = re.sub(r'/\*.*?\*/', '', value, flags=re.DOTALL)
    # 4. Remove common hex/unicode escape prefixes (evasion tricks)
    value = value.replace('\\x', '').replace('\\u', '')
    # 5. Normalize excessive whitespace
    value = ' '.join(value.split())
    return value


def _check_string_value(value: str, column: str, row_idx: str) -> list[ContentFinding]:
    """Apply deobfuscation then all text rules against a single string cell value."""
    deobfuscated = _deobfuscate(value)
    findings = []
    for rule_id, severity, category, description, pattern in _RULES:
        # Check both original and deobfuscated value for coverage
        match = pattern.search(deobfuscated) or pattern.search(value)
        if match:
            target = deobfuscated if pattern.search(deobfuscated) else value
            sample = target[max(0, match.start() - 20): match.start() + _SAMPLE_MAX_LEN]
            findings.append(ContentFinding(
                rule_id=rule_id,
                severity=severity,
                category=category,
                description=description,
                location=f"column={column!r}, row={row_idx}",
                sample=sample[:_SAMPLE_MAX_LEN],
            ))
    return findings


def _load_dataframe(path: Path) -> tuple[pd.DataFrame | None, str | None]:
    """
    Load dataset into DataFrame. Returns (df, error_msg).
    Supports CSV, Parquet, JSON, JSONL, XLSX (openpyxl read_only), TXT.
    """
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            df = pd.read_csv(path, nrows=_MAX_ROWS, low_memory=False, dtype=str)

        elif suffix == ".parquet":
            # Stream row groups to avoid loading entire file into memory
            try:
                import pyarrow.parquet as pq
                pf = pq.ParquetFile(path)
                rows_collected: list[pd.DataFrame] = []
                rows_read = 0
                for batch in pf.iter_batches(batch_size=2000):
                    chunk_df = batch.to_pandas().astype(str)
                    rows_remaining = _MAX_ROWS - rows_read
                    rows_collected.append(chunk_df.head(rows_remaining))
                    rows_read += len(chunk_df)
                    if rows_read >= _MAX_ROWS:
                        break
                df = pd.concat(rows_collected, ignore_index=True) if rows_collected else pd.DataFrame()
            except ImportError:
                # Fallback: pandas read with head truncation
                df = pd.read_parquet(path)
                df = df.head(_MAX_ROWS).astype(str)

        elif suffix == ".json":
            with path.open("r", encoding="utf-8", errors="replace") as fh:
                raw = json.load(fh)
            if isinstance(raw, list):
                df = pd.DataFrame(raw[:_MAX_ROWS]).astype(str)
            elif isinstance(raw, dict):
                df = pd.DataFrame([raw]).astype(str)
            else:
                return None, f"Unsupported JSON root type: {type(raw).__name__}"

        elif suffix == ".jsonl":
            df = pd.read_json(io.StringIO(path.read_text("utf-8", errors="replace")),
                              lines=True, nrows=_MAX_ROWS, dtype=str)

        elif suffix in (".xlsx", ".xls"):
            # Check file size before loading (openpyxl loads full sheet)
            size_mb = path.stat().st_size / (1024 * 1024)
            if size_mb > _XLSX_MAX_MB:
                return None, f"Excel file too large ({size_mb:.1f} MB). Max {_XLSX_MAX_MB} MB for XLSX scanning."
            # Use openpyxl read_only mode for memory efficiency
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=str(path), read_only=True, data_only=True)
                ws = wb.active
                rows = []
                headers: list[str] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                    else:
                        rows.append([str(c) if c is not None else "" for c in row])
                        if len(rows) >= _MAX_ROWS:
                            break
                wb.close()
                df = pd.DataFrame(rows, columns=headers[:len(rows[0])] if rows else headers)
            except ImportError:
                df = pd.read_excel(path, engine="openpyxl", nrows=_MAX_ROWS, dtype=str)

        elif suffix == ".txt":
            # Scan as single-column text file line by line
            lines = []
            with path.open("r", encoding="utf-8", errors="replace") as fh:
                for i, line in enumerate(fh):
                    if i >= _TXT_MAX_LINES:
                        break
                    lines.append({"text": line.rstrip("\n")})
            df = pd.DataFrame(lines)

        else:
            return None, f"Unsupported file format: {suffix!r}"

        return df, None
    except Exception as exc:  # noqa: BLE001
        return None, f"Failed to parse dataset: {exc}"


def raw_bytes_scan(path: Path) -> list[ContentFinding]:
    """
    Stage 0: Scan raw file bytes BEFORE parsing as a dataset.
    Catches binary malware (EICAR, PE/ELF executables, shellcode) that would be
    lost or corrupted when the file is parsed through pandas/openpyxl.

    This is critical: a CSV containing an EICAR string or embedded PE header
    must be caught at the binary level, not after text-field parsing.
    """
    findings = []
    try:
        raw = path.read_bytes()
    except Exception as exc:
        logger.warning("raw_bytes_scan: could not read %s — %s", path.name, exc)
        return findings

    # ── EICAR test string (exact bytes) ──────────────────────────────────────
    if _EICAR in raw:
        findings.append(ContentFinding(
            rule_id="MAL-001",
            severity="critical",
            category="malware_signature",
            description="EICAR antivirus test string detected in raw file bytes",
            location="raw_bytes",
            sample="X5O!P%@AP[4\\PZX54(P^)7CC)7}$EICAR...",
        ))

    # ── Windows PE / MZ executable header ────────────────────────────────────
    if raw[:2] == b"MZ" or b"MZ\x90\x00" in raw[:4096] or b"MZP\x00" in raw[:4096]:
        findings.append(ContentFinding(
            rule_id="MAL-002",
            severity="critical",
            category="malware_signature",
            description="Windows PE/MZ executable header found in raw file bytes",
            location="raw_bytes:offset=0",
            sample="MZ (PE executable header)",
        ))

    # ── ELF Linux executable header ───────────────────────────────────────────
    if raw[:4] == b"\x7fELF":
        findings.append(ContentFinding(
            rule_id="MAL-003",
            severity="critical",
            category="malware_signature",
            description="ELF Linux/Unix executable binary detected in raw file bytes",
            location="raw_bytes:offset=0",
            sample="\\x7fELF (ELF binary header)",
        ))

    # ── Embedded PE header anywhere in file (polyglot files) ─────────────────
    mz_offset = raw.find(b"MZ", 512)   # A-010: .find() returns -1, never raises ValueError
    if raw[:2] != b"MZ" and mz_offset != -1:
        findings.append(ContentFinding(
            rule_id="MAL-002",
            severity="critical",
            category="malware_signature",
            description=f"Embedded Windows PE/MZ executable found at offset {mz_offset} (polyglot file)",
            location=f"raw_bytes:offset={mz_offset}",
            sample=f"MZ header at byte {mz_offset}",
        ))

    # ── Shellcode NOP sled pattern (repeated 0x90 bytes) ─────────────────────
    if b"\x90\x90\x90\x90\x90\x90\x90\x90\x90\x90\x90\x90" in raw:
        findings.append(ContentFinding(
            rule_id="MAL-011",
            severity="high",
            category="shellcode",
            description="NOP sled pattern detected in raw bytes (classic shellcode technique)",
            location="raw_bytes",
            sample="\\x90 * 12+ (NOP sled)",
        ))

    logger.info(
        "raw_bytes_scan: %s — %d bytes, %d raw findings",
        path.name, len(raw), len(findings),
    )
    return findings


def check_file(file_path: str) -> ContentCheckResult:
    """
    Entry point: inspect dataset for threat patterns.

    Pipeline:
      Stage 0 — Raw bytes scan (EICAR, PE/ELF headers, shellcode) — runs FIRST
      Stage 1 — Dataframe content inspection (injection, script, SQL, malware names)

    Returns a ContentCheckResult with findings, row count, and risk score.
    """
    path = Path(file_path)
    result = ContentCheckResult()

    # ── Stage 0: Raw bytes scan (catches binary malware before parsing) ───────
    raw_findings = raw_bytes_scan(path)
    result.findings.extend(raw_findings)
    if raw_findings:
        logger.warning(
            "Stage 0 raw scan: %d binary threat(s) found in %s",
            len(raw_findings), path.name,
        )

    # ── Stage 1: Dataframe content inspection ─────────────────────────────────
    df, error = _load_dataframe(path)
    if error:
        result.error = error
        logger.warning("Content check parse error for %s: %s", file_path, error)
        return result

    result.rows_inspected = len(df)
    seen_rules_per_column: set[tuple[str, str]] = set()

    # Pre-seed seen set with rules already triggered by raw scan
    for f in raw_findings:
        seen_rules_per_column.add((f.rule_id, "raw_bytes"))

    for col in df.columns:
        try:
            col_values = df[col].dropna().astype(str)
        except Exception:  # noqa: BLE001
            continue

        for row_idx, value in col_values.items():
            for finding in _check_string_value(value, str(col), str(row_idx)):  # type: ignore[arg-type]
                dedup_key = (finding.rule_id, str(col))
                if dedup_key not in seen_rules_per_column:
                    seen_rules_per_column.add(dedup_key)
                    result.findings.append(finding)

    return result

